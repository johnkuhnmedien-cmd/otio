"""Research-Excel-Parser + Dateinamen für Adobe-Stock-Import."""

from __future__ import annotations

import json
from pathlib import Path

import pytest

from otio_app.services.adobe_research_import import (
    STATUS_DOWNLOADED,
    STATUS_OPEN,
    build_research_import_board,
    cleanup_media_folder_json,
    download_research_import,
    format_asset_stem,
    parse_research_excel,
    sanitize_folder_name,
)

IRELAND_XLSX = (
    Path(__file__).resolve().parent / "fixtures" / "research_template_ireland.xlsx"
)


def test_sanitize_folder_name_keeps_readable_title() -> None:
    assert sanitize_folder_name("Dublin") == "Dublin"
    assert "Giant" in sanitize_folder_name("Giant’s Causeway & Causeway Coast")
    assert "/" not in sanitize_folder_name("A/B:C")


def test_format_asset_stem() -> None:
    assert format_asset_stem("Dublin", 1) == "Dublin_Asset_01"
    assert format_asset_stem("Dublin", 12) == "Dublin_Asset_12"


@pytest.mark.skipif(not IRELAND_XLSX.is_file(), reason="Ireland research Excel fehlt")
def test_parse_ireland_research_excel() -> None:
    plan = parse_research_excel(IRELAND_XLSX)
    assert plan.chapter_count >= 8
    assert plan.asset_count >= 100
    titles = [ch.title for ch in plan.chapters]
    assert "Dublin" in titles
    assert "Skellig Michael" in titles
    dublin = next(ch for ch in plan.chapters if ch.title == "Dublin")
    assert dublin.folder_name == "Dublin"
    assert dublin.asset_count >= 20
    assert dublin.assets[0].asset_id.isdigit()
    assert "stock.adobe.com" in dublin.assets[0].link
    # Non-numeric note in ID column must be skipped (Dingle row anomaly).
    dingle = next(ch for ch in plan.chapters if "Dingle" in ch.title)
    assert all(a.asset_id.isdigit() for a in dingle.assets)


def test_parse_minimal_workbook(tmp_path: Path) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Alpha"
    ws["D1"] = "Beta"
    ws["A2"] = "Count"
    ws["B2"] = "Asset ID"
    ws["C2"] = "Link"
    ws["D2"] = "Count"
    ws["E2"] = "Asset ID"
    ws["F2"] = "Link"
    ws["A3"] = 1
    ws["B3"] = 123456
    ws["C3"] = "https://stock.adobe.com/ph/video/x/123456"
    ws["D3"] = 1
    ws["E3"] = 987654
    ws["F3"] = "https://stock.adobe.com/ph/images/y/987654"
    path = tmp_path / "mini.xlsx"
    wb.save(path)

    plan = parse_research_excel(path)
    assert plan.chapter_count == 2
    assert plan.chapters[0].title == "Alpha"
    assert plan.chapters[0].assets[0].asset_id == "123456"
    assert plan.chapters[0].assets[0].media_hint == "video"
    assert plan.chapters[1].assets[0].media_hint == "image"


def _minimal_plan(tmp_path: Path):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Alpha"
    ws["A2"] = "Count"
    ws["B2"] = "Asset ID"
    ws["C2"] = "Link"
    ws["A3"] = 1
    ws["B3"] = 111
    ws["C3"] = "https://stock.adobe.com/ph/video/x/111"
    ws["A4"] = 2
    ws["B4"] = 222
    ws["C4"] = "https://stock.adobe.com/ph/video/x/222"
    path = tmp_path / "board.xlsx"
    wb.save(path)
    return parse_research_excel(path)


def test_build_research_import_board_marks_downloaded_from_state_manifest(
    tmp_path: Path,
) -> None:
    plan = _minimal_plan(tmp_path)
    target = tmp_path / "out"
    state = tmp_path / "state"
    state.mkdir()
    media = target / "Alpha" / "Alpha_Asset_01.mp4"
    media.parent.mkdir(parents=True)
    media.write_bytes(b"x" * 1000)
    (state / "adobe_research_import_manifest.json").write_text(
        json.dumps(
            {
                "items": [
                    {
                        "asset_id": "111",
                        "status": "downloaded",
                        "license": "Video_HD",
                        "local_path": str(media),
                    }
                ]
            }
        ),
        encoding="utf-8",
    )

    board = build_research_import_board(plan, target, state_dir=state)
    assert board.total == 2
    assert board.downloaded == 1
    assert board.open_count == 1
    by_id = {a.asset_id: a for ch in board.chapters for a in ch.assets}
    assert by_id["111"].status == STATUS_DOWNLOADED
    assert by_id["222"].status == STATUS_OPEN


def test_cleanup_media_folder_json_removes_sidecars_and_root_json(tmp_path: Path) -> None:
    target = tmp_path / "media"
    chapter = target / "Alpha"
    chapter.mkdir(parents=True)
    (chapter / "Alpha_Asset_01.adobe.json").write_text("{}", encoding="utf-8")
    (target / "adobe_research_import_board.json").write_text("{}", encoding="utf-8")
    (target / "adobe_research_import_manifest.json").write_text("{}", encoding="utf-8")
    removed = cleanup_media_folder_json(target)
    assert removed["sidecar"] == 1
    assert removed["board"] == 1
    assert removed["manifest"] == 1
    assert not list(target.rglob("*.json"))


def test_video_license_path_uses_files_meta_and_video_licenses(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """Videos: Files-API → Video_4K/HD; Comp/cancelled zählt nicht als Erfolg."""
    from otio_app.services import adobe_research_import as mod

    monkeypatch.setattr(mod, "_ASSET_PAUSE_SECONDS", 0)
    monkeypatch.setattr(mod, "_LICENSE_RETRY_PAUSE_SECONDS", 0)

    calls: list[str] = []

    class _Adapter:
        def lookup_file_metadata(self, content_id, api_key):
            return {"media_type_id": 4, "content_type": "video/mp4"}

        def content_info_purchase(self, content_id, license_type, api_key, access_token):
            return {"state": "not_purchased"}

        def find_license_history_download(self, content_id, api_key, access_token, *, pages=5):
            return None

        def _license_asset(self, content_id, license_type, api_key, access_token, *, diagnose=True):
            calls.append(license_type)
            raise RuntimeError(
                f"Adobe-Lizenzierung nicht bestätigt: state=cancelled size=Comp license={license_type}"
            )

        def _stream_download_to_file(self, *args, **kwargs):
            raise AssertionError("kein Download erwartet")

    monkeypatch.setattr(mod, "get_api_key", lambda key: "x")
    monkeypatch.setattr(mod, "get_adobe_access_token", lambda: "tok")

    with pytest.raises(RuntimeError, match="cct_pro_unlimited_images|Video_4K|cancelled"):
        mod._license_and_download_to_path(
            _Adapter(),
            content_id="644202290",
            media_type="video",
            destination=tmp_path / "x",
            media_hint="video",
        )
    assert calls == ["Video_4K", "Video_HD"]


def test_photo_uses_standard_license(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    from otio_app.services import adobe_research_import as mod

    monkeypatch.setattr(mod, "_ASSET_PAUSE_SECONDS", 0)
    monkeypatch.setattr(mod, "_LICENSE_RETRY_PAUSE_SECONDS", 0)
    calls: list[str] = []

    class _Adapter:
        def lookup_file_metadata(self, content_id, api_key):
            return {"media_type_id": 1, "content_type": "image/jpeg"}

        def content_info_purchase(self, *_a, **_k):
            return {"state": "not_purchased"}

        def find_license_history_download(self, *_a, **_k):
            return None

        def _license_asset(self, content_id, license_type, api_key, access_token, *, diagnose=True):
            calls.append(license_type)
            return {
                "state": "just_purchased",
                "license": license_type,
                "url": f"https://stock.adobe.com/Rest/Libraries/Download/{content_id}/1",
                "content_type": "image/jpeg",
            }

        def _stream_download_to_file(self, url, local_path, *, api_key, access_token, size, max_bytes):
            local_path.write_bytes(b"x" * 200_000)

    monkeypatch.setattr(mod, "get_api_key", lambda key: "x")
    monkeypatch.setattr(mod, "get_adobe_access_token", lambda: "tok")

    path, license_type = mod._license_and_download_to_path(
        _Adapter(),
        content_id="111",
        media_type="image",
        destination=tmp_path / "photo",
        media_hint="image",
    )
    assert license_type.startswith("Standard")
    assert calls == ["Standard"]
    assert path.is_file()


def test_4k_over_600mb_falls_back_to_hd(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    from otio_app.services import adobe_research_import as mod
    from otio_app.services.supplement_sources.adobe_stock import AdobeAssetTooLargeError

    monkeypatch.setattr(mod, "_ASSET_PAUSE_SECONDS", 0)
    monkeypatch.setattr(mod, "_LICENSE_RETRY_PAUSE_SECONDS", 0)
    downloads: list[tuple[int | None, int | None]] = []

    class _Adapter:
        def lookup_file_metadata(self, content_id, api_key):
            return {"media_type_id": 4, "content_type": "video/mp4"}

        def content_info_purchase(self, *_a, **_k):
            return {"state": "not_purchased"}

        def find_license_history_download(self, *_a, **_k):
            return None

        def _license_asset(self, content_id, license_type, api_key, access_token, *, diagnose=True):
            return {
                "state": "just_purchased",
                "license": license_type,
                "url": f"https://stock.adobe.com/Rest/Libraries/Download/{content_id}/4",
                "content_type": "video/mp4",
            }

        def _stream_download_to_file(self, url, local_path, *, api_key, access_token, size, max_bytes):
            downloads.append((size, max_bytes))
            if max_bytes is not None:
                raise AdobeAssetTooLargeError()
            local_path.write_bytes(b"x" * 200_000)

    monkeypatch.setattr(mod, "get_api_key", lambda key: "x")
    monkeypatch.setattr(mod, "get_adobe_access_token", lambda: "tok")

    path, license_type = mod._license_and_download_to_path(
        _Adapter(),
        content_id="999",
        media_type="video",
        destination=tmp_path / "big",
        media_hint="video",
    )
    assert license_type.startswith("Video_HD (4K>600MB)")
    assert "MB" in license_type
    assert downloads[0][0] == 2160 and downloads[0][1] == mod.ADOBE_STOCK_VIDEO_4K_MAX_BYTES
    assert downloads[1][0] == 1080 and downloads[1][1] is None
    assert path.is_file()


def test_content_info_4k_over_600mb_continues_to_hd(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """Bereits gekauftes 4K >600MB darf nicht als Erfolg enden — Fallback HD."""
    from otio_app.services import adobe_research_import as mod
    from otio_app.services.supplement_sources.adobe_stock import AdobeAssetTooLargeError

    monkeypatch.setattr(mod, "_ASSET_PAUSE_SECONDS", 0)
    monkeypatch.setattr(mod, "_LICENSE_RETRY_PAUSE_SECONDS", 0)
    license_calls: list[str] = []

    class _Adapter:
        def lookup_file_metadata(self, content_id, api_key):
            return {"media_type_id": 4, "content_type": "video/mp4"}

        def content_info_purchase(self, content_id, license_type, api_key, access_token):
            if license_type == "Video_4K":
                return {
                    "state": "purchased",
                    "license": "Video_4K",
                    "url": f"https://stock.adobe.com/Rest/Libraries/Download/{content_id}/4",
                    "content_type": "video/mp4",
                }
            return {"state": "not_purchased"}

        def find_license_history_download(self, *_a, **_k):
            return None

        def _license_asset(self, content_id, license_type, api_key, access_token, *, diagnose=True):
            license_calls.append(license_type)
            return {
                "state": "just_purchased",
                "license": license_type,
                "url": f"https://stock.adobe.com/Rest/Libraries/Download/{content_id}/hd",
                "content_type": "video/mp4",
            }

        def _stream_download_to_file(self, url, local_path, *, api_key, access_token, size, max_bytes):
            if max_bytes is not None:
                raise AdobeAssetTooLargeError()
            local_path.write_bytes(b"x" * 200_000)

    monkeypatch.setattr(mod, "get_api_key", lambda key: "x")
    monkeypatch.setattr(mod, "get_adobe_access_token", lambda: "tok")

    path, license_type = mod._license_and_download_to_path(
        _Adapter(),
        content_id="777",
        media_type="video",
        destination=tmp_path / "from-info",
        media_hint="video",
    )
    assert license_type.startswith("Video_HD (4K>600MB)")
    assert license_calls == ["Video_HD"]
    assert path.is_file()


def test_post_download_size_guard_rejects_oversized_4k(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """Wenn der Stream die Grenze verfehlt, greift der Dateigrößen-Check."""
    from otio_app.services import adobe_research_import as mod

    monkeypatch.setattr(mod, "_ASSET_PAUSE_SECONDS", 0)
    monkeypatch.setattr(mod, "_LICENSE_RETRY_PAUSE_SECONDS", 0)
    # Kleine Grenze (über MIN_DOWNLOAD), damit der Test keine 600-MB-Datei schreibt.
    monkeypatch.setattr(mod, "ADOBE_STOCK_VIDEO_4K_MAX_BYTES", 150_000)

    class _Adapter:
        def lookup_file_metadata(self, content_id, api_key):
            return {"media_type_id": 4, "content_type": "video/mp4"}

        def content_info_purchase(self, *_a, **_k):
            return {"state": "not_purchased"}

        def find_license_history_download(self, *_a, **_k):
            return None

        def _license_asset(self, content_id, license_type, api_key, access_token, *, diagnose=True):
            return {
                "state": "just_purchased",
                "license": license_type,
                "url": f"https://stock.adobe.com/Rest/Libraries/Download/{content_id}/4",
                "content_type": "video/mp4",
            }

        def _stream_download_to_file(self, url, local_path, *, api_key, access_token, size, max_bytes):
            # Simuliert: Stream-Limit greift nicht, Datei ist trotzdem zu groß.
            if max_bytes is not None:
                local_path.write_bytes(b"x" * (max_bytes + 1024))
            else:
                local_path.write_bytes(b"x" * 200_000)

    monkeypatch.setattr(mod, "get_api_key", lambda key: "x")
    monkeypatch.setattr(mod, "get_adobe_access_token", lambda: "tok")

    path, license_type = mod._license_and_download_to_path(
        _Adapter(),
        content_id="888",
        media_type="video",
        destination=tmp_path / "guard",
        media_hint="video",
    )
    assert license_type.startswith("Video_HD (4K>600MB)")
    assert path.stat().st_size == 200_000


def test_already_licensed_uses_content_info_url(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    from otio_app.services import adobe_research_import as mod

    monkeypatch.setattr(mod, "_ASSET_PAUSE_SECONDS", 0)
    monkeypatch.setattr(mod, "_LICENSE_RETRY_PAUSE_SECONDS", 0)

    class _Adapter:
        def lookup_file_metadata(self, content_id, api_key):
            return {"media_type_id": 4, "content_type": "video/mp4"}

        def content_info_purchase(self, content_id, license_type, api_key, access_token):
            if license_type == "Video_HD":
                return {
                    "state": "purchased",
                    "license": "Video_HD",
                    "url": f"https://stock.adobe.com/Rest/Libraries/Download/{content_id}/4",
                    "content_type": "video/mp4",
                }
            return {"state": "not_purchased"}

        def find_license_history_download(self, *_a, **_k):
            return None

        def _license_asset(self, content_id, license_type, api_key, access_token, *, diagnose=True):
            # 4K nicht lizenziert / nicht vorhanden — HD kommt aus Content/Info.
            raise RuntimeError(f"cancelled {license_type}")

        def _stream_download_to_file(self, url, local_path, *, api_key, access_token, size, max_bytes):
            local_path.write_bytes(b"x" * 200_000)

    monkeypatch.setattr(mod, "get_api_key", lambda key: "x")
    monkeypatch.setattr(mod, "get_adobe_access_token", lambda: "tok")

    path, license_type = mod._license_and_download_to_path(
        _Adapter(),
        content_id="222",
        media_type="video",
        destination=tmp_path / "vid",
        media_hint="video",
    )
    assert license_type.startswith("Video_HD")
    assert path.suffix == ".mp4"


def test_download_respects_should_stop(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    from otio_app.services import adobe_research_import as mod

    monkeypatch.setattr(mod, "_ASSET_PAUSE_SECONDS", 0)
    monkeypatch.setattr(mod, "_LICENSE_RETRY_PAUSE_SECONDS", 0)

    plan = _minimal_plan(tmp_path)
    target = tmp_path / "dl"
    calls = {"n": 0}

    def fake_license_and_download(adapter, *, content_id, media_type, destination, media_hint=""):
        calls["n"] += 1
        path = destination.with_suffix(".mp4")
        path.write_bytes(b"x" * 200_000)
        return path, "Video_HD"

    monkeypatch.setattr(
        "otio_app.services.adobe_research_import._license_and_download_to_path",
        fake_license_and_download,
    )

    class _Ready:
        acquire_enabled = True
        message = "ok"

    monkeypatch.setattr(
        "otio_app.services.adobe_research_import.AdobeStockAdapter.readiness",
        lambda self: _Ready(),
    )

    stop_after_first = {"stop": False}

    def should_stop() -> bool:
        return stop_after_first["stop"]

    progress_events = []

    def on_progress(event) -> None:
        progress_events.append(event)
        if event.done >= 1 and event.status == "downloaded":
            stop_after_first["stop"] = True

    state = tmp_path / "state"
    result = download_research_import(
        plan,
        target,
        state_dir=state,
        skip_existing_ids=False,
        progress_callback=on_progress,
        should_stop=should_stop,
    )
    assert result.cancelled is True
    assert result.downloaded == 1
    assert calls["n"] == 1
    assert any(e.status == "cancelled" for e in progress_events)
    assert not list(target.rglob("*.json"))
    assert (state / "adobe_research_import_manifest.json").is_file()
    board = build_research_import_board(plan, target, state_dir=state)
    assert board.downloaded == 1
    assert board.open_count >= 1
