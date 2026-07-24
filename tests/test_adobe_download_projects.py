"""Persistente Adobe Download-Projekte."""

from __future__ import annotations

from pathlib import Path

import pytest

from otio_app.services import adobe_download_projects as store


@pytest.fixture()
def projects_home(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> Path:
    root = tmp_path / "data"
    root.mkdir()
    monkeypatch.setattr(store, "ensure_data_dir", lambda: root)
    return root


def _excel_bytes() -> bytes:
    from io import BytesIO

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Alpha"
    ws["A2"] = "Count"
    ws["B2"] = "Asset ID"
    ws["C2"] = "Link"
    ws["A3"] = 1
    ws["B3"] = 111
    ws["C3"] = "https://stock.adobe.com/ph/video/x/111"
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_create_list_switch_delete(projects_home: Path, tmp_path: Path) -> None:
    target_a = tmp_path / "Ireland"
    target_b = tmp_path / "Scotland"
    a = store.create_download_project(
        name="Irland",
        target_root=str(target_a),
        excel_bytes=_excel_bytes(),
        excel_filename="ireland.xlsx",
    )
    b = store.create_download_project(
        name="Schottland",
        target_root=str(target_b),
        excel_bytes=_excel_bytes(),
        excel_filename="scotland.xlsx",
    )
    listed = store.list_download_projects()
    assert {p.id for p in listed} == {a.id, b.id}
    assert store.project_excel_path(a.id).is_file()

    plan = store.load_project_plan(a.id)
    assert plan.chapter_count == 1
    assert plan.chapters[0].title == "Alpha"

    updated = store.update_download_project(a.id, name="Irland v2", selected_chapters=["Alpha"])
    assert updated.name == "Irland v2"
    assert updated.selected_chapters == ["Alpha"]

    store.delete_download_project(a.id)
    assert store.get_download_project(a.id) is None
    assert store.get_download_project(b.id) is not None
