"""Research-Excel → Adobe-Stock lizenzieren/herunterladen (vor Projektanlage).

Erwartetes Excel-Layout (Research Template):
- Zeile 1: Kapitel-Überschriften alle 3 Spalten (1, 4, 7, …)
- Zeile 2: Count | Asset ID | Link (wiederholt)
- ab Zeile 3: laufende Nummer | Adobe Content-ID | Adobe-URL
"""

from __future__ import annotations

import json
import re
import time
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import BinaryIO, Callable, Iterable

from otio_app.defaults import (
    ADOBE_STOCK_LICENSE_TYPE_STANDARD,
    ADOBE_STOCK_LICENSE_TYPE_VIDEO_4K,
    ADOBE_STOCK_LICENSE_TYPE_VIDEO_HD,
    ADOBE_STOCK_MIN_DOWNLOAD_BYTES,
    ADOBE_STOCK_VIDEO_4K_MAX_BYTES,
)
from otio_app.services.adobe_stock_oauth import decode_access_token_claims, get_adobe_access_token
from otio_app.services.api_keys import get_api_key
from otio_app.services.media_utils import probe_duration_seconds
from otio_app.services.supplement_sources.adobe_stock import (
    VIDEO_ENTITLEMENT_HINT,
    AdobeAssetTooLargeError,
    AdobeAuthenticationExpiredError,
    AdobeContentUnavailableError,
    AdobeIdentityChangedError,
    AdobeImportError,
    AdobeLicenseNotPossibleError,
    AdobeLicenseTransactionCancelledError,
    AdobeRateLimitedError,
    AdobeStockAdapter,
    AdobeVideoEntitlementError,
    AdobeWatermarkedPreviewError,
    DownloadedMediaInvalidError,
    LocalStorageError,
    classify_adobe_url,
    is_full_adobe_download_url,
    token_fingerprint,
)

__all__ = [
    "AdobeResearchAsset",
    "AdobeResearchAssetStatus",
    "AdobeResearchChapter",
    "AdobeResearchChapterStatus",
    "AdobeResearchImportBoard",
    "AdobeResearchImportPlan",
    "AdobeResearchImportProgress",
    "AdobeResearchImportResult",
    "build_research_import_board",
    "cleanup_media_folder_json",
    "download_research_import",
    "format_asset_stem",
    "parse_research_excel",
    "persist_research_import_board",
    "sanitize_folder_name",
]

_INVALID_FS_CHARS = re.compile(r'[<>:"/\\|?*\x00-\x1f]')
_WHITESPACE = re.compile(r"\s+")
_MANIFEST_NAME = "adobe_research_import_manifest.json"
_BOARD_NAME = "adobe_research_import_board.json"
# Bewusst langsam / strikt sequenziell — ein Asset nach dem anderen.
# Zu schnelle API-/Download-Ketten führen bei Adobe oft zu Comp/cancelled.
_ASSET_PAUSE_SECONDS = 2.5  # vor Start jedes Assets (Files/Info/License)
_API_CALL_PAUSE_SECONDS = 1.0  # vor jedem Adobe-API-Call
_LICENSE_RETRY_PAUSE_SECONDS = 1.2  # zwischen Lizenzvarianten (4K → HD)
_DOWNLOAD_START_PAUSE_SECONDS = 1.0  # nach Lizenz-URL, vor Stream-Start
_POST_ASSET_PAUSE_SECONDS = 2.0  # nach Umbenennen, bevor das nächste Asset startet


def _sleep(seconds: float) -> None:
    if seconds and seconds > 0:
        time.sleep(seconds)


def _redact_email(email: str) -> str:
    text = (email or "").strip()
    if not text or "@" not in text:
        return text[:3] + "…" if text else ""
    local, _, domain = text.partition("@")
    keep = local[:2] if len(local) > 2 else local[:1]
    return f"{keep}…@{domain}"

STATUS_DOWNLOADED = "downloaded"
STATUS_OPEN = "open"
STATUS_ERROR = "error"
STATUS_DOWNLOADING = "downloading"
STATUS_CANCELLED = "cancelled"
STATUS_SKIPPED = "skipped"
STATUS_UNAVAILABLE = "unavailable"


@dataclass(frozen=True)
class AdobeResearchAsset:
    asset_id: str
    link: str = ""
    media_hint: str = ""  # "video" | "image" | ""
    row_number: int = 0
    column_block: int = 0


@dataclass(frozen=True)
class AdobeResearchChapter:
    title: str
    folder_name: str
    assets: tuple[AdobeResearchAsset, ...] = ()
    source_column: int = 1

    @property
    def asset_count(self) -> int:
        return len(self.assets)


@dataclass(frozen=True)
class AdobeResearchImportPlan:
    sheet_name: str
    chapters: tuple[AdobeResearchChapter, ...] = ()

    @property
    def chapter_count(self) -> int:
        return len(self.chapters)

    @property
    def asset_count(self) -> int:
        return sum(ch.asset_count for ch in self.chapters)


@dataclass
class AdobeResearchImportItemResult:
    chapter_title: str
    folder_name: str
    asset_id: str
    status: str  # downloaded | skipped | error | cancelled
    local_path: str = ""
    message: str = ""
    license: str = ""


@dataclass
class AdobeResearchImportResult:
    target_root: str
    items: list[AdobeResearchImportItemResult] = field(default_factory=list)
    manifest_path: str = ""
    cancelled: bool = False
    diagnostics: dict = field(default_factory=dict)

    @property
    def downloaded(self) -> int:
        return sum(1 for item in self.items if item.status == STATUS_DOWNLOADED)

    @property
    def skipped(self) -> int:
        return sum(1 for item in self.items if item.status == STATUS_SKIPPED)

    @property
    def errors(self) -> int:
        return sum(1 for item in self.items if item.status == STATUS_ERROR)

    @property
    def unavailable(self) -> int:
        return sum(1 for item in self.items if item.status == STATUS_UNAVAILABLE)

    @property
    def cancelled_count(self) -> int:
        return sum(1 for item in self.items if item.status == STATUS_CANCELLED)


@dataclass(frozen=True)
class AdobeResearchAssetStatus:
    chapter_title: str
    folder_name: str
    asset_id: str
    link: str = ""
    status: str = STATUS_OPEN
    local_path: str = ""
    license: str = ""
    message: str = ""


@dataclass(frozen=True)
class AdobeResearchChapterStatus:
    title: str
    folder_name: str
    assets: tuple[AdobeResearchAssetStatus, ...] = ()

    @property
    def total(self) -> int:
        return len(self.assets)

    @property
    def downloaded(self) -> int:
        return sum(1 for a in self.assets if a.status == STATUS_DOWNLOADED)

    @property
    def open_count(self) -> int:
        return sum(1 for a in self.assets if a.status in {STATUS_OPEN, STATUS_CANCELLED})

    @property
    def error_count(self) -> int:
        return sum(1 for a in self.assets if a.status == STATUS_ERROR)

    @property
    def downloading_count(self) -> int:
        return sum(1 for a in self.assets if a.status == STATUS_DOWNLOADING)


@dataclass(frozen=True)
class AdobeResearchImportBoard:
    sheet_name: str
    target_root: str
    chapters: tuple[AdobeResearchChapterStatus, ...] = ()

    @property
    def total(self) -> int:
        return sum(ch.total for ch in self.chapters)

    @property
    def downloaded(self) -> int:
        return sum(ch.downloaded for ch in self.chapters)

    @property
    def open_count(self) -> int:
        return sum(ch.open_count for ch in self.chapters)

    @property
    def error_count(self) -> int:
        return sum(ch.error_count for ch in self.chapters)


@dataclass(frozen=True)
class AdobeResearchImportProgress:
    done: int
    total: int
    folder_name: str
    asset_id: str
    chapter_title: str = ""
    status: str = STATUS_DOWNLOADING
    message: str = ""
    fraction: float = 0.0


def sanitize_folder_name(title: str) -> str:
    """Dateisystemtauglicher Ordnername aus Excel-Überschrift."""
    text = (title or "").strip()
    text = text.replace("’", "'").replace("‘", "'").replace("–", "-").replace("—", "-")
    text = _INVALID_FS_CHARS.sub("-", text)
    text = _WHITESPACE.sub(" ", text).strip(" .")
    return text or "Untitled"


def format_asset_stem(folder_name: str, index: int) -> str:
    """z. B. Dublin_Asset_01"""
    return f"{folder_name}_Asset_{index:02d}"


def _as_asset_id(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return str(value) if value > 0 else None
    if isinstance(value, float):
        if value.is_integer() and value > 0:
            return str(int(value))
        return None
    text = str(value).strip()
    if not text:
        return None
    if text.isdigit():
        return text
    # Excel manchmal "1.23456789E9"
    try:
        as_float = float(text.replace(",", ""))
    except ValueError:
        return None
    if as_float.is_integer() and as_float > 0:
        return str(int(as_float))
    return None


def _media_hint_from_link(link: str) -> str:
    lower = (link or "").lower()
    if "/images/" in lower or "/image/" in lower:
        return "image"
    if "/video/" in lower:
        return "video"
    return ""


def parse_research_excel(
    source: str | Path | bytes | BinaryIO,
    *,
    sheet_name: str | None = None,
) -> AdobeResearchImportPlan:
    """Parst Research-Template-Excel zu Kapitelblöcken."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError(
            "Paket 'openpyxl' fehlt — bitte `pip install openpyxl` ausführen."
        ) from exc

    if isinstance(source, bytes):
        from io import BytesIO

        source = BytesIO(source)
    wb = load_workbook(source, data_only=True, read_only=True)
    try:
        name = sheet_name or wb.sheetnames[0]
        if name not in wb.sheetnames:
            raise ValueError(f"Sheet '{name}' nicht gefunden. Vorhanden: {wb.sheetnames}")
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if len(rows) < 3:
        raise ValueError("Excel zu kurz — erwartet Zeile 1 Titel, Zeile 2 Header, ab Zeile 3 IDs.")

    header_row = rows[0]
    max_col = len(header_row)
    chapters: list[AdobeResearchChapter] = []

    col = 0  # 0-based
    while col < max_col:
        title_raw = header_row[col] if col < len(header_row) else None
        title = str(title_raw).strip() if title_raw is not None else ""
        if not title:
            col += 3
            continue

        assets: list[AdobeResearchAsset] = []
        seen_ids: set[str] = set()
        for row_index, row in enumerate(rows[2:], start=3):
            # Pad short rows
            id_col = col + 1
            link_col = col + 2
            asset_val = row[id_col] if id_col < len(row) else None
            link_val = row[link_col] if link_col < len(row) else None
            asset_id = _as_asset_id(asset_val)
            link = str(link_val).strip() if link_val is not None else ""
            if asset_id is None:
                # Manchmal steht Text in der ID-Spalte (Notiz) — überspringen.
                continue
            if asset_id in seen_ids:
                continue
            seen_ids.add(asset_id)
            assets.append(
                AdobeResearchAsset(
                    asset_id=asset_id,
                    link=link,
                    media_hint=_media_hint_from_link(link),
                    row_number=row_index,
                    column_block=col + 1,
                )
            )

        chapters.append(
            AdobeResearchChapter(
                title=title,
                folder_name=sanitize_folder_name(title),
                assets=tuple(assets),
                source_column=col + 1,
            )
        )
        col += 3

    if not chapters:
        raise ValueError(
            "Keine Kapitel gefunden. Erwartet: Überschrift in Zeile 1 alle 3 Spalten, "
            "Asset-IDs in der 2. Spalte jedes Blocks."
        )

    return AdobeResearchImportPlan(sheet_name=name, chapters=tuple(chapters))


def _manifest_records(manifest_path: Path) -> dict[str, dict]:
    if not manifest_path.is_file():
        return {}
    try:
        payload = json.loads(manifest_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    found: dict[str, dict] = {}
    for item in payload.get("items") or []:
        aid = str(item.get("asset_id") or "").strip()
        if not aid:
            continue
        status = str(item.get("status") or STATUS_OPEN)
        if status == STATUS_SKIPPED:
            status = STATUS_DOWNLOADED
        found[aid] = {
            "local_path": str(item.get("local_path") or ""),
            "license": str(item.get("license") or ""),
            "message": str(item.get("message") or ""),
            "status": status,
        }
    return found


def _legacy_sidecar_records(target_root: Path) -> dict[str, dict]:
    """Einmalige Migration: alte *.adobe.json neben Medien lesen."""
    found: dict[str, dict] = {}
    if not target_root.is_dir():
        return found
    for path in target_root.rglob("*.adobe.json"):
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue
        aid = str(payload.get("asset_id") or "").strip()
        if not aid:
            continue
        found[aid] = {
            "local_path": str(payload.get("local_path") or ""),
            "license": str(payload.get("license") or ""),
            "message": "",
            "status": STATUS_DOWNLOADED,
        }
    return found


def _downloaded_ids_from_records(records: dict[str, dict]) -> set[str]:
    ids: set[str] = set()
    for aid, record in records.items():
        status = str(record.get("status") or "")
        if status not in {STATUS_DOWNLOADED, STATUS_SKIPPED}:
            continue
        local = str(record.get("local_path") or "")
        if local and not Path(local).is_file():
            # Datei fehlt → als noch offen behandeln
            continue
        ids.add(aid)
    return ids


def cleanup_media_folder_json(target_root: str | Path) -> dict[str, int]:
    """Löscht Board/Manifest/*.adobe.json aus dem Medien-Zielordner (nicht aus data/)."""
    root = Path(target_root).expanduser().resolve()
    removed = {"sidecar": 0, "board": 0, "manifest": 0}
    if not root.is_dir():
        return removed
    board = root / _BOARD_NAME
    if board.is_file():
        board.unlink()
        removed["board"] = 1
    manifest = root / _MANIFEST_NAME
    if manifest.is_file():
        manifest.unlink()
        removed["manifest"] = 1
    for path in root.rglob("*.adobe.json"):
        try:
            path.unlink()
            removed["sidecar"] += 1
        except OSError:
            continue
    return removed


def build_research_import_board(
    plan: AdobeResearchImportPlan,
    target_root: str | Path | None,
    *,
    state_dir: str | Path | None = None,
    live_statuses: dict[str, dict] | None = None,
) -> AdobeResearchImportBoard:
    """Excel-Spiegel: pro Asset Downloaded / Open / Error (+ Live-Overrides).

    Fortschritt wird aus `state_dir` gelesen (Download-Projekt unter data/),
    nicht aus JSON neben den Mediendateien. Legacy-Sidecars im Zielordner
    werden nur noch als Fallback gelesen.
    """
    root = Path(target_root).expanduser().resolve() if target_root else Path()
    state = Path(state_dir).expanduser().resolve() if state_dir else None
    state_records = _manifest_records(state / _MANIFEST_NAME) if state else {}
    # Legacy: alte Dateien im Medienordner (Migration)
    legacy_root = _manifest_records(root / _MANIFEST_NAME) if target_root else {}
    legacy_sidecars = _legacy_sidecar_records(root) if target_root else {}
    live = live_statuses or {}

    chapters: list[AdobeResearchChapterStatus] = []
    for chapter in plan.chapters:
        assets: list[AdobeResearchAssetStatus] = []
        for asset in chapter.assets:
            record = (
                live.get(asset.asset_id)
                or state_records.get(asset.asset_id)
                or legacy_sidecars.get(asset.asset_id)
                or legacy_root.get(asset.asset_id)
            )
            if record:
                status = str(record.get("status") or STATUS_DOWNLOADED)
                if status == STATUS_SKIPPED:
                    status = STATUS_DOWNLOADED
                local_path = str(record.get("local_path") or "")
                if (
                    status == STATUS_DOWNLOADED
                    and local_path
                    and not Path(local_path).is_file()
                ):
                    status = STATUS_OPEN
                assets.append(
                    AdobeResearchAssetStatus(
                        chapter_title=chapter.title,
                        folder_name=chapter.folder_name,
                        asset_id=asset.asset_id,
                        link=asset.link,
                        status=status,
                        local_path=local_path if status == STATUS_DOWNLOADED else "",
                        license=str(record.get("license") or ""),
                        message=str(record.get("message") or ""),
                    )
                )
            else:
                assets.append(
                    AdobeResearchAssetStatus(
                        chapter_title=chapter.title,
                        folder_name=chapter.folder_name,
                        asset_id=asset.asset_id,
                        link=asset.link,
                        status=STATUS_OPEN,
                    )
                )
        chapters.append(
            AdobeResearchChapterStatus(
                title=chapter.title,
                folder_name=chapter.folder_name,
                assets=tuple(assets),
            )
        )
    return AdobeResearchImportBoard(
        sheet_name=plan.sheet_name,
        target_root=str(root) if target_root else "",
        chapters=tuple(chapters),
    )


def persist_research_import_board(
    board: AdobeResearchImportBoard,
    *,
    state_dir: str | Path | None = None,
) -> Path | None:
    """Schreibt Board-JSON nur in state_dir (nie in den Medienordner)."""
    if state_dir is None:
        return None
    root = Path(state_dir).expanduser().resolve()
    root.mkdir(parents=True, exist_ok=True)
    payload = {
        "schema_version": "adobe-research-import-board-v1",
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "sheet_name": board.sheet_name,
        "target_root": board.target_root,
        "downloaded": board.downloaded,
        "open": board.open_count,
        "errors": board.error_count,
        "total": board.total,
        "chapters": [
            {
                "title": ch.title,
                "folder_name": ch.folder_name,
                "downloaded": ch.downloaded,
                "open": ch.open_count,
                "errors": ch.error_count,
                "total": ch.total,
                "assets": [asdict(a) for a in ch.assets],
            }
            for ch in board.chapters
        ],
    }
    path = root / _BOARD_NAME
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def _next_asset_index(folder: Path) -> int:
    """Nächster freier Asset_NN Index in einem Kapitelordner."""
    pattern = re.compile(r"_Asset_(\d+)\.[^.]+$", re.IGNORECASE)
    highest = 0
    if folder.is_dir():
        for path in folder.iterdir():
            if not path.is_file():
                continue
            # Unfertige Downloads (.part) nicht als fertige Assets zählen.
            if path.name.endswith(".part"):
                continue
            match = pattern.search(path.name)
            if match:
                highest = max(highest, int(match.group(1)))
    return highest + 1


def _media_type_from_file_meta(meta: dict, *, hint: str = "") -> str:
    if hint in {"video", "image"}:
        return hint
    media_type_id = meta.get("media_type_id")
    try:
        media_type_id_i = int(media_type_id) if media_type_id is not None else 0
    except (TypeError, ValueError):
        media_type_id_i = 0
    from otio_app.defaults import ADOBE_STOCK_MEDIA_TYPE_ID_PHOTO, ADOBE_STOCK_MEDIA_TYPE_ID_VIDEO

    if media_type_id_i == ADOBE_STOCK_MEDIA_TYPE_ID_VIDEO:
        return "video"
    if media_type_id_i == ADOBE_STOCK_MEDIA_TYPE_ID_PHOTO:
        return "image"
    content_type = str(meta.get("content_type") or "").lower()
    if content_type.startswith("video/"):
        return "video"
    if content_type.startswith("image/"):
        return "image"
    return "video"


def _extension_for_purchase(purchase: dict, media_type: str) -> str:
    content_type = str(purchase.get("content_type") or "").lower()
    if "video" in content_type:
        return ".mp4"
    if "png" in content_type:
        return ".png"
    if "jpeg" in content_type or "jpg" in content_type:
        return ".jpg"
    return ".jpg" if media_type == "image" else ".mp4"


def _format_license_with_size(license_type: str, local_path: Path) -> str:
    """Lizenzlabel inkl. Dateigröße für Live-Log / Board."""
    try:
        mb = local_path.stat().st_size / (1024 * 1024)
    except OSError:
        return license_type
    return f"{license_type} · {mb:.0f} MB"


def _download_purchase_to_path(
    adapter: AdobeStockAdapter,
    purchase: dict,
    destination: Path,
    *,
    api_key: str,
    access_token: str,
    media_type: str,
    size: int | None,
    max_bytes: int | None,
    phase_callback: Callable[[str], None] | None = None,
) -> Path:
    """Download zuerst als `.part`, danach atomar auf den Zielnamen umbenennen."""
    url = str(purchase.get("url") or "")
    url_class = classify_adobe_url(url)
    if url_class == "watermarked":
        adapter.request_counters.watermarked += 1
        raise AdobeWatermarkedPreviewError(
            "Watermarked-URL wird nicht als Vollversion gespeichert "
            f"(url_class={url_class})."
        )
    if url_class != "download" or not is_full_adobe_download_url(url):
        raise RuntimeError(
            f"Keine Voll-Download-URL (url_class={url_class}). "
            "Vollständige Download-URLs werden nicht protokolliert."
        )
    final_path = destination.with_suffix(_extension_for_purchase(purchase, media_type))
    part_path = final_path.with_name(final_path.name + ".part")
    part_path.unlink(missing_ok=True)
    final_path.unlink(missing_ok=True)

    if phase_callback is not None:
        phase_callback("warte vor Download…")
    _sleep(_DOWNLOAD_START_PAUSE_SECONDS)
    if phase_callback is not None:
        phase_callback("Download läuft…")
    try:
        adapter._stream_download_to_file(
            url,
            part_path,
            api_key=api_key,
            access_token=access_token,
            size=size,
            max_bytes=max_bytes,
        )
        if not part_path.is_file() or part_path.stat().st_size < ADOBE_STOCK_MIN_DOWNLOAD_BYTES:
            part_path.unlink(missing_ok=True)
            adapter.request_counters.invalid_media += 1
            raise DownloadedMediaInvalidError(
                f"Download zu klein / ungültig: {final_path.name}",
                details={"path": str(final_path)},
            )
        # Sicherheitsnetz: auch wenn Content-Length fehlte / Stream-Abbruch
        # nicht greift, darf 4K die 600-MB-Grenze nicht behalten.
        if max_bytes is not None and part_path.stat().st_size > max_bytes:
            part_path.unlink(missing_ok=True)
            raise AdobeAssetTooLargeError(
                f"Download {final_path.name} überschreitet "
                f"{max_bytes / (1024 * 1024):.0f} MB-Grenze."
            )
        if media_type == "video":
            duration = probe_duration_seconds(part_path)
            if duration is None:
                part_path.unlink(missing_ok=True)
                adapter.request_counters.invalid_media += 1
                raise DownloadedMediaInvalidError(
                    f"Lokale Videodatei technisch ungültig: {final_path.name}",
                    details={"path": str(final_path)},
                )
        if phase_callback is not None:
            phase_callback(f"umbenennen → {final_path.name}")
        try:
            part_path.replace(final_path)
        except OSError as exc:
            part_path.unlink(missing_ok=True)
            adapter.request_counters.local_storage_errors += 1
            raise LocalStorageError(
                f"Umbenennen fehlgeschlagen: {exc}",
                details={"path": str(final_path)},
            ) from exc
        return final_path
    except AdobeImportError:
        part_path.unlink(missing_ok=True)
        raise
    except Exception:
        part_path.unlink(missing_ok=True)
        raise


def _license_and_download_to_path(
    adapter: AdobeStockAdapter,
    *,
    content_id: str,
    media_type: str,
    destination: Path,
    media_hint: str = "",
    phase_callback: Callable[[str], None] | None = None,
) -> tuple[Path, str]:
    """Lizenziert/lädt eine Content-ID — Hot-Path ohne LicenseHistory-Vollscan.

    Reihenfolge:
    1) Files-API → Medientyp
    2) Content/Info (bereits purchased + Voll-URL)
    3) Content/License (license_again nie true)
    4) Video_4K → bei TooLarge oder cancelled: begrenzter Video_HD-Fallback
    """
    def _phase(message: str) -> None:
        if phase_callback is not None:
            phase_callback(message)

    api_key = get_api_key("ADOBE_STOCK_API_KEY")
    access_token = get_adobe_access_token()
    if not api_key:
        raise PermissionError("ADOBE_STOCK_API_KEY fehlt.")
    if not access_token:
        raise PermissionError(
            "Kein Adobe Access-Token — bitte OAuth-Login nutzen oder "
            "ADOBE_STOCK_ACCESS_TOKEN setzen."
        )

    destination.parent.mkdir(parents=True, exist_ok=True)

    _phase("warte vor Files-API…")
    _sleep(_API_CALL_PAUSE_SECONDS)
    _phase("Files-API (Medientyp)…")
    try:
        meta = adapter.lookup_file_metadata(content_id, api_key)
        resolved_type = _media_type_from_file_meta(meta, hint=media_hint or media_type)
    except AdobeContentUnavailableError:
        raise
    except Exception:
        resolved_type = media_hint if media_hint in {"video", "image"} else media_type
        meta = {}

    if resolved_type == "image":
        pending: list[tuple[str, int | None, int | None]] = [
            (ADOBE_STOCK_LICENSE_TYPE_STANDARD, None, None)
        ]
    else:
        pending = [
            (
                ADOBE_STOCK_LICENSE_TYPE_VIDEO_4K,
                2160,
                ADOBE_STOCK_VIDEO_4K_MAX_BYTES,
            ),
            (ADOBE_STOCK_LICENSE_TYPE_VIDEO_HD, 1080, None),
        ]

    attempt_errors: list[str] = []
    cancelled_licenses: set[str] = set()
    tried: set[str] = set()
    while pending:
        license_type, size, max_bytes = pending.pop(0)
        if license_type in tried:
            continue
        tried.add(license_type)
        _phase(f"warte vor {license_type}…")
        _sleep(_LICENSE_RETRY_PAUSE_SECONDS if len(tried) > 1 else _API_CALL_PAUSE_SECONDS)

        # Bereits lizenziert? Content/Info zuerst (kein neuer Kauf / kein license_again).
        _phase(f"Content/Info ({license_type})…")
        info = adapter.content_info_purchase(
            content_id, license_type, api_key, access_token
        )
        info_state = str(info.get("state") or "")
        info_url = str(info.get("url") or "")
        if info_state in {"purchased", "just_purchased"} and is_full_adobe_download_url(info_url):
            try:
                path = _download_purchase_to_path(
                    adapter,
                    info,
                    destination,
                    api_key=api_key,
                    access_token=access_token,
                    media_type=resolved_type,
                    size=size,
                    max_bytes=max_bytes,
                    phase_callback=phase_callback,
                )
                used = str(info.get("license") or license_type)
                if max_bytes is not None and any(">600MB" in e for e in attempt_errors):
                    used = f"{used} (nach 4K>600MB)"
                return path, _format_license_with_size(used, path)
            except AdobeAssetTooLargeError:
                attempt_errors.append(f"{license_type}: >600MB → Fallback HD")
                continue
            except (LocalStorageError, DownloadedMediaInvalidError, AdobeRateLimitedError):
                raise
            except AdobeImportError as exc:
                attempt_errors.append(f"{license_type}: Info-Download [{exc.code}] {exc}")
            except Exception as exc:  # noqa: BLE001
                attempt_errors.append(f"{license_type}: Info-Download {exc}")

        _phase(f"warte vor License ({license_type})…")
        _sleep(_API_CALL_PAUSE_SECONDS)
        _phase(f"License ({license_type})…")
        try:
            purchase = adapter._license_asset(
                content_id,
                license_type,
                api_key,
                access_token,
                diagnose=False,
            )
        except AdobeContentUnavailableError:
            raise
        except AdobeRateLimitedError:
            raise
        except AdobeVideoEntitlementError:
            raise
        except AdobeAuthenticationExpiredError:
            # Ein Refresh-Versuch, dann erneut — sonst stoppen.
            try:
                access_token = get_adobe_access_token(force_refresh=True) or access_token
                purchase = adapter._license_asset(
                    content_id,
                    license_type,
                    api_key,
                    access_token,
                    diagnose=False,
                )
            except AdobeAuthenticationExpiredError:
                raise
            except AdobeVideoEntitlementError:
                raise
            except AdobeImportError as exc:
                attempt_errors.append(f"{license_type}: [{exc.code}] {exc}")
                if isinstance(exc, AdobeLicenseTransactionCancelledError):
                    cancelled_licenses.add(license_type)
                continue
        except AdobeLicenseTransactionCancelledError as exc:
            cancelled_licenses.add(license_type)
            attempt_errors.append(f"{license_type}: [{exc.code}] {exc}")
            # 4K cancelled → HD versuchen (bestehendes begrenztes Fallback)
            continue
        except (AdobeLicenseNotPossibleError, AdobeWatermarkedPreviewError, AdobeImportError) as exc:
            attempt_errors.append(f"{license_type}: [{getattr(exc, 'code', 'error')}] {exc}")
            continue
        except Exception as exc:  # noqa: BLE001
            msg = str(exc)
            if "does not match type of content" in msg:
                if resolved_type == "video" and ADOBE_STOCK_LICENSE_TYPE_STANDARD not in tried:
                    pending.append((ADOBE_STOCK_LICENSE_TYPE_STANDARD, None, None))
                if resolved_type == "image":
                    if ADOBE_STOCK_LICENSE_TYPE_VIDEO_4K not in tried:
                        pending.append(
                            (
                                ADOBE_STOCK_LICENSE_TYPE_VIDEO_4K,
                                2160,
                                ADOBE_STOCK_VIDEO_4K_MAX_BYTES,
                            )
                        )
                    if ADOBE_STOCK_LICENSE_TYPE_VIDEO_HD not in tried:
                        pending.append((ADOBE_STOCK_LICENSE_TYPE_VIDEO_HD, 1080, None))
            attempt_errors.append(f"{license_type}: {exc}")
            continue

        try:
            path = _download_purchase_to_path(
                adapter,
                purchase,
                destination,
                api_key=api_key,
                access_token=access_token,
                media_type=resolved_type,
                size=size,
                max_bytes=max_bytes,
                phase_callback=phase_callback,
            )
            used = license_type
            if license_type == ADOBE_STOCK_LICENSE_TYPE_VIDEO_HD and any(
                ">600MB" in e for e in attempt_errors
            ):
                used = "Video_HD (4K>600MB)"
            return path, _format_license_with_size(used, path)
        except AdobeAssetTooLargeError:
            attempt_errors.append(f"{license_type}: >600MB → Fallback HD")
            continue
        except (LocalStorageError, DownloadedMediaInvalidError, AdobeRateLimitedError):
            raise
        except AdobeImportError as exc:
            attempt_errors.append(f"{license_type}: Download [{exc.code}] {exc}")
            continue
        except Exception as exc:  # noqa: BLE001
            attempt_errors.append(f"{license_type}: Download {exc}")
            continue

    # Beide Video-Lizenzen cancelled → Batch soll stoppen (Caller wertet Code aus).
    if (
        ADOBE_STOCK_LICENSE_TYPE_VIDEO_4K in cancelled_licenses
        and ADOBE_STOCK_LICENSE_TYPE_VIDEO_HD in cancelled_licenses
    ):
        raise AdobeLicenseTransactionCancelledError(
            f"Video_4K und Video_HD cancelled für Content-ID {content_id}. "
            f"Batch wird kontrolliert gestoppt. Versuche: {' | '.join(attempt_errors)}"
        )

    detail = " | ".join(attempt_errors) if attempt_errors else "unbekannter Fehler"
    raise RuntimeError(
        f"Adobe-Download fehlgeschlagen für Content-ID {content_id} "
        f"(media_type={resolved_type}, meta_content_type={meta.get('content_type') or '—'}). "
        f"Versuche: {detail}. {VIDEO_ENTITLEMENT_HINT}"
    )


def download_research_import(
    plan: AdobeResearchImportPlan,
    target_root: str | Path,
    *,
    state_dir: str | Path | None = None,
    chapter_titles: Iterable[str] | None = None,
    skip_existing_ids: bool = True,
    progress_callback: Callable[[AdobeResearchImportProgress], None] | None = None,
    should_stop: Callable[[], bool] | None = None,
    live_status_callback: Callable[[dict[str, dict]], None] | None = None,
) -> AdobeResearchImportResult:
    """Lizenzieren + Download in Zielordner/{Kapitel}/{Kapitel}_Asset_NN.ext.

    Strikt sequenziell pro Asset: Pause → API → Download als `.part` →
    Umbenennen auf Endnamen → Pause. Fortschritt/Manifest landen in
    `state_dir` (Download-Projekt unter data/), nicht als JSON neben den
    Medien. `should_stop` bricht kooperativ zwischen Assets ab.
    """
    root = Path(target_root).expanduser().resolve()
    root.mkdir(parents=True, exist_ok=True)
    state_path = Path(state_dir).expanduser().resolve() if state_dir else None
    if state_path is not None:
        state_path.mkdir(parents=True, exist_ok=True)

    selected = None
    if chapter_titles is not None:
        selected = {str(t).strip() for t in chapter_titles if str(t).strip()}

    adapter = AdobeStockAdapter()
    adapter.reset_request_diagnostics()
    readiness = adapter.readiness()
    if not readiness.acquire_enabled:
        raise PermissionError(
            readiness.message
            or "Adobe Stock ist nicht für Lizenzierung/Download konfiguriert."
        )

    result = AdobeResearchImportResult(target_root=str(root))
    chapters = [
        ch
        for ch in plan.chapters
        if selected is None or ch.title in selected or ch.folder_name in selected
    ]
    total = sum(ch.asset_count for ch in chapters)
    done = 0
    live_statuses: dict[str, dict] = {}
    stopped = False
    batch_stop_reason = ""

    # OAuth-Identität am Batch-Start — Wechsel mittendrin stoppt den Batch.
    start_claims = decode_access_token_claims()
    start_sub = str(start_claims.get("sub") or "")
    start_token = get_adobe_access_token() or ""
    start_fp = token_fingerprint(start_token)
    result.diagnostics = {
        "batch_id": adapter.batch_id,
        "oauth_sub": start_sub,
        "oauth_email_redacted": _redact_email(str(start_claims.get("email") or "")),
        "token_fingerprint": start_fp,
        "request_counters": adapter.request_counters.as_dict(),
    }

    prior_records: dict[str, dict] = {}
    if state_path is not None:
        prior_records.update(_manifest_records(state_path / _MANIFEST_NAME))
    prior_records.update(_legacy_sidecar_records(root))
    prior_records.update(_manifest_records(root / _MANIFEST_NAME))
    already_global = _downloaded_ids_from_records(prior_records) if skip_existing_ids else set()

    def _emit(
        *,
        folder_name: str,
        asset_id: str,
        chapter_title: str,
        status: str,
        message: str = "",
    ) -> None:
        if progress_callback is None:
            return
        fraction = (done / total) if total else 1.0
        progress_callback(
            AdobeResearchImportProgress(
                done=done,
                total=total,
                folder_name=folder_name,
                asset_id=asset_id,
                chapter_title=chapter_title,
                status=status,
                message=message,
                fraction=min(1.0, max(0.0, fraction)),
            )
        )

    def _publish_live() -> None:
        if live_status_callback is not None:
            live_status_callback(dict(live_statuses))
        board = build_research_import_board(
            plan,
            root,
            state_dir=state_path,
            live_statuses=live_statuses,
        )
        persist_research_import_board(board, state_dir=state_path)

    for chapter in chapters:
        if stopped:
            break
        folder = root / chapter.folder_name
        folder.mkdir(parents=True, exist_ok=True)
        already = set(already_global)
        next_index = _next_asset_index(folder)

        for asset in chapter.assets:
            # Identitätswechsel während des Batches → sofort stoppen.
            now_claims = decode_access_token_claims()
            now_sub = str(now_claims.get("sub") or "")
            if start_sub and now_sub and now_sub != start_sub:
                stopped = True
                batch_stop_reason = "adobe_identity_changed"
                msg = (
                    f"[{AdobeIdentityChangedError.code}] OAuth-sub wechselte "
                    f"während des Batches ({start_sub[:8]}… → {now_sub[:8]}…)."
                )
                live_statuses[asset.asset_id] = {
                    "status": STATUS_ERROR,
                    "message": msg,
                    "local_path": "",
                    "license": "",
                }
                result.items.append(
                    AdobeResearchImportItemResult(
                        chapter_title=chapter.title,
                        folder_name=chapter.folder_name,
                        asset_id=asset.asset_id,
                        status=STATUS_ERROR,
                        message=msg,
                    )
                )
                _publish_live()
                break

            if should_stop is not None and should_stop():
                stopped = True
                result.cancelled = True
                live_statuses[asset.asset_id] = {
                    "status": STATUS_CANCELLED,
                    "message": "Import gestoppt — noch offen",
                    "local_path": "",
                    "license": "",
                }
                result.items.append(
                    AdobeResearchImportItemResult(
                        chapter_title=chapter.title,
                        folder_name=chapter.folder_name,
                        asset_id=asset.asset_id,
                        status=STATUS_CANCELLED,
                        message="Import gestoppt — noch offen",
                    )
                )
                _publish_live()
                _emit(
                    folder_name=chapter.folder_name,
                    asset_id=asset.asset_id,
                    chapter_title=chapter.title,
                    status=STATUS_CANCELLED,
                    message="Import gestoppt",
                )
                break

            done += 1
            if asset.asset_id in already:
                live_statuses[asset.asset_id] = {
                    "status": STATUS_DOWNLOADED,
                    "message": "bereits vorhanden",
                    "local_path": "",
                    "license": "",
                }
                result.items.append(
                    AdobeResearchImportItemResult(
                        chapter_title=chapter.title,
                        folder_name=chapter.folder_name,
                        asset_id=asset.asset_id,
                        status=STATUS_SKIPPED,
                        message="bereits vorhanden",
                    )
                )
                _publish_live()
                _emit(
                    folder_name=chapter.folder_name,
                    asset_id=asset.asset_id,
                    chapter_title=chapter.title,
                    status=STATUS_SKIPPED,
                    message="bereits vorhanden",
                )
                continue

            # Ein Asset nach dem anderen: Pause → History/API → Download(.part) → Rename → Pause.
            def _phase(message: str) -> None:
                live_statuses[asset.asset_id] = {
                    "status": STATUS_DOWNLOADING,
                    "message": message,
                    "local_path": "",
                    "license": "",
                }
                _publish_live()
                _emit(
                    folder_name=chapter.folder_name,
                    asset_id=asset.asset_id,
                    chapter_title=chapter.title,
                    status=STATUS_DOWNLOADING,
                    message=message,
                )

            _phase("warte vor nächstem Asset…")
            _sleep(_ASSET_PAUSE_SECONDS)

            stem = format_asset_stem(chapter.folder_name, next_index)
            dest = folder / stem  # Suffix setzt Download/Rename
            adapter.asset_index = done
            try:
                local_path, used_license = _license_and_download_to_path(
                    adapter,
                    content_id=asset.asset_id,
                    media_type=asset.media_hint or "video",
                    destination=dest,
                    media_hint=asset.media_hint,
                    phase_callback=_phase,
                )
                live_statuses[asset.asset_id] = {
                    "status": STATUS_DOWNLOADED,
                    "message": local_path.name,
                    "local_path": str(local_path),
                    "license": used_license,
                }
                result.items.append(
                    AdobeResearchImportItemResult(
                        chapter_title=chapter.title,
                        folder_name=chapter.folder_name,
                        asset_id=asset.asset_id,
                        status=STATUS_DOWNLOADED,
                        local_path=str(local_path),
                        license=used_license,
                        message=local_path.name,
                    )
                )
                already.add(asset.asset_id)
                already_global.add(asset.asset_id)
                next_index += 1
                _publish_live()
                _emit(
                    folder_name=chapter.folder_name,
                    asset_id=asset.asset_id,
                    chapter_title=chapter.title,
                    status=STATUS_DOWNLOADED,
                    message=f"{used_license} → {local_path.name}",
                )
                _sleep(_POST_ASSET_PAUSE_SECONDS)
            except AdobeContentUnavailableError as exc:
                live_statuses[asset.asset_id] = {
                    "status": STATUS_UNAVAILABLE,
                    "message": str(exc),
                    "local_path": "",
                    "license": "",
                }
                result.items.append(
                    AdobeResearchImportItemResult(
                        chapter_title=chapter.title,
                        folder_name=chapter.folder_name,
                        asset_id=asset.asset_id,
                        status=STATUS_UNAVAILABLE,
                        message=str(exc),
                    )
                )
                _publish_live()
                _emit(
                    folder_name=chapter.folder_name,
                    asset_id=asset.asset_id,
                    chapter_title=chapter.title,
                    status=STATUS_UNAVAILABLE,
                    message=str(exc),
                )
            except AdobeImportError as exc:
                code = getattr(exc, "code", "adobe_import_error")
                details = getattr(exc, "details", {}) or {}
                req_id = details.get("request_id") or ""
                msg = f"[{code}] {exc}"
                if req_id:
                    msg = f"{msg} (X-Request-Id={req_id})"
                live_statuses[asset.asset_id] = {
                    "status": STATUS_ERROR,
                    "message": msg,
                    "local_path": "",
                    "license": "",
                }
                result.items.append(
                    AdobeResearchImportItemResult(
                        chapter_title=chapter.title,
                        folder_name=chapter.folder_name,
                        asset_id=asset.asset_id,
                        status=STATUS_ERROR,
                        message=msg,
                    )
                )
                _publish_live()
                _emit(
                    folder_name=chapter.folder_name,
                    asset_id=asset.asset_id,
                    chapter_title=chapter.title,
                    status=STATUS_ERROR,
                    message=msg,
                )
                # Rate-Limit erschöpft / 4K+HD cancelled / Identity → Batch stoppen
                if code in {
                    "adobe_rate_limited",
                    "adobe_license_transaction_cancelled",
                    "adobe_identity_changed",
                    "adobe_authentication_expired",
                }:
                    # cancelled nur stoppen wenn beide Video-Lizenzen betroffen
                    if code != "adobe_license_transaction_cancelled" or (
                        "Video_4K und Video_HD cancelled" in str(exc)
                    ):
                        stopped = True
                        batch_stop_reason = code
                        break
            except Exception as exc:  # noqa: BLE001
                live_statuses[asset.asset_id] = {
                    "status": STATUS_ERROR,
                    "message": str(exc),
                    "local_path": "",
                    "license": "",
                }
                result.items.append(
                    AdobeResearchImportItemResult(
                        chapter_title=chapter.title,
                        folder_name=chapter.folder_name,
                        asset_id=asset.asset_id,
                        status=STATUS_ERROR,
                        message=str(exc),
                    )
                )
                _publish_live()
                _emit(
                    folder_name=chapter.folder_name,
                    asset_id=asset.asset_id,
                    chapter_title=chapter.title,
                    status=STATUS_ERROR,
                    message=str(exc),
                )

    board = build_research_import_board(
        plan,
        root,
        state_dir=state_path,
        live_statuses=live_statuses,
    )
    persist_research_import_board(board, state_dir=state_path)

    # Manifest mergen: vorherige Downloads behalten + aktuelle Run-Items
    merged_by_id: dict[str, dict] = dict(prior_records)
    for item in result.items:
        merged_by_id[item.asset_id] = asdict(item)
    manifest_items = list(merged_by_id.values())
    manifest = {
        "schema_version": "adobe-research-import-v1",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "target_root": str(root),
        "sheet_name": plan.sheet_name,
        "downloaded": sum(
            1
            for item in manifest_items
            if str(item.get("status")) in {STATUS_DOWNLOADED, STATUS_SKIPPED}
        ),
        "skipped": result.skipped,
        "errors": result.errors,
        "cancelled": result.cancelled,
        "items": manifest_items,
    }
    if state_path is not None:
        manifest_path = state_path / _MANIFEST_NAME
        manifest_path.write_text(
            json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        result.manifest_path = str(manifest_path)
    else:
        result.manifest_path = ""

    # Redigierte Request-Diagnose (keine Tokens/URLs)
    recent_events = [
        e.as_dict()
        for e in adapter.request_diag_events[-40:]
    ]
    result.diagnostics = {
        "batch_id": adapter.batch_id,
        "oauth_sub": start_sub,
        "oauth_email_redacted": _redact_email(str(start_claims.get("email") or "")),
        "token_fingerprint": start_fp,
        "batch_stop_reason": batch_stop_reason,
        "request_counters": adapter.request_counters.as_dict(),
        "recent_requests": recent_events,
    }
    if state_path is not None:
        diag_path = state_path / "adobe_research_import_diagnostics.json"
        diag_path.write_text(
            json.dumps(result.diagnostics, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
    return result
